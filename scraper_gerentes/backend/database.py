"""
Capa de base de datos — SQLite por defecto, compatible con MySQL.
"""

import sqlite3, os, logging
from datetime import datetime
from typing import List, Optional, Dict

logger = logging.getLogger(__name__)

DB_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "gerentes.db")


def get_connection():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True) if os.path.dirname(DB_PATH) else None
    with get_connection() as conn:
        # Comprobar si la tabla existe y tiene la columna fuente
        tabla_existe = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='empresas'"
        ).fetchone()
        
        if tabla_existe:
            columnas = [row[1] for row in conn.execute("PRAGMA table_info(empresas)").fetchall()]
            if "fuente" not in columnas:
                # BD antigua sin columna fuente — añadirla
                try:
                    conn.execute("ALTER TABLE empresas ADD COLUMN fuente TEXT")
                    conn.commit()
                    logger.info("Migración: columna 'fuente' añadida.")
                except Exception as e:
                    logger.error(f"Error en migración: {e}")

        conn.executescript("""
            CREATE TABLE IF NOT EXISTS empresas (
                id               INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre           TEXT NOT NULL,
                provincia        TEXT,
                email            TEXT,
                telefono         TEXT,
                web              TEXT,
                direccion        TEXT,
                facturacion_eur  REAL,
                cnae_cpv         TEXT,
                url_datoscif     TEXT,
                fuente           TEXT,
                gerente          TEXT,
                estado           TEXT DEFAULT 'pendiente',
                error_msg        TEXT,
                fecha_scraping   TIMESTAMP,
                fecha_carga      TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE INDEX IF NOT EXISTS idx_nombre ON empresas(nombre);
            CREATE INDEX IF NOT EXISTS idx_estado  ON empresas(estado);
        """)
    logger.info("Base de datos inicializada.")

def cargar_empresas_desde_excel(filepath: str) -> int:
    import pandas as pd
    df = pd.read_excel(filepath)
    df = df.where(pd.notna(df), None)
    inserted = 0
    with get_connection() as conn:
        for _, row in df.iterrows():
            nombre = str(row.get("nombre", "") or "").strip()
            if not nombre:
                continue
            exists = conn.execute("SELECT id FROM empresas WHERE nombre = ?", (nombre,)).fetchone()
            if exists:
                continue
            conn.execute("""
                INSERT INTO empresas (nombre, provincia, email, telefono, web, direccion, facturacion_eur, cnae_cpv)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                nombre,
                row.get("provincia"),
                row.get("email"),
                row.get("telefono"),
                row.get("web"),
                row.get("direccion"),
                row.get("facturacion_eur"),
                str(row.get("cnae_cpv", "") or ""),
            ))
            inserted += 1
    logger.info(f"Insertadas {inserted} empresas nuevas.")
    return inserted


def get_empresas(estado: Optional[str] = None, limit: int = 100, offset: int = 0) -> List[Dict]:
    with get_connection() as conn:
        if estado:
            rows = conn.execute(
                "SELECT * FROM empresas WHERE estado = ? ORDER BY id LIMIT ? OFFSET ?",
                (estado, limit, offset)
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM empresas ORDER BY id LIMIT ? OFFSET ?",
                (limit, offset)
            ).fetchall()
        return [dict(r) for r in rows]


def get_empresas_pendientes() -> List[Dict]:
    with get_connection() as conn:
        rows = conn.execute(
            "SELECT * FROM empresas WHERE estado = 'pendiente' ORDER BY id"
        ).fetchall()
        return [dict(r) for r in rows]


def actualizar_empresa(empresa_id: int, gerente: Optional[str], url: Optional[str],
                       estado: str, fuente: Optional[str] = None, error: Optional[str] = None):
    with get_connection() as conn:
        conn.execute("""
            UPDATE empresas
            SET gerente=?, url_datoscif=?, fuente=?, estado=?, error_msg=?, fecha_scraping=?
            WHERE id=?
        """, (gerente, url, fuente, estado, error, datetime.now().isoformat(), empresa_id))


def get_stats() -> Dict:
    with get_connection() as conn:
        row = conn.execute("""
            SELECT
                COUNT(*) as total,
                SUM(CASE WHEN estado='pendiente'    THEN 1 ELSE 0 END) as pendientes,
                SUM(CASE WHEN estado='ok'           THEN 1 ELSE 0 END) as con_gerente,
                SUM(CASE WHEN estado='sin_gerente'  THEN 1 ELSE 0 END) as sin_gerente,
                SUM(CASE WHEN estado='no_encontrada'THEN 1 ELSE 0 END) as no_encontradas,
                SUM(CASE WHEN estado='error'        THEN 1 ELSE 0 END) as errores
            FROM empresas
        """).fetchone()
        return dict(row)


def exportar_a_excel(output_path: str) -> str:
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl import load_workbook

    with get_connection() as conn:
        rows = conn.execute("""
            SELECT nombre, provincia, gerente, fuente, email, telefono, web,
                   direccion, facturacion_eur, url_datoscif, estado, fecha_scraping
            FROM empresas ORDER BY nombre
        """).fetchall()

    df = pd.DataFrame([dict(r) for r in rows])
    df.columns = ["Empresa","Provincia","Gerente","Fuente","Email","Teléfono","Web",
                  "Dirección","Facturación (€)","URL Ficha","Estado","Fecha Scraping"]
    df.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.active
    hf = PatternFill("solid", fgColor="1B4F72")
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    ok_fill = PatternFill("solid", fgColor="D5F5E3")
    for cell in ws[1]:
        cell.fill = hf; cell.font = hfont
        cell.alignment = Alignment(horizontal="center")
    for row in ws.iter_rows(min_row=2):
        if row[10].value == "ok":
            for cell in row: cell.fill = ok_fill
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(w+4, 60)
    wb.save(output_path)
    return output_path
